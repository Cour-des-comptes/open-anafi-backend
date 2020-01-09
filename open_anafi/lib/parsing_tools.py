import openpyxl
import re
from django.db import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from open_anafi.models import Variable, Indicator, IndicatorParameter, User, IndicatorLibelle
from open_anafi.serializers import VariableSerializer, IndicatorSerializer
from .ply.parser_anafi import parse_equation
from .ply.parsing_classes import Var, Indic
from .indicator_tools import IndicatorParameterTools, IndicatorTools
from .frame_tools import FrameTools
import queue
import logging

logger = logging.getLogger(__name__)


class ExcelEntry:
    def __init__(self, name, libelle, formula, description):
        self.name = name
        self.libelle = libelle
        self.formula = formula
        self.description = description


class ContainsUndefinedIndicators(Exception):
    pass


def create_variable(variable, indicator_parameter):
    """
	Create an open_anafi variable object linked to an indicator parameter

	:param variable: The variable to create
	:type variable: class:`open_anafi.lib.ply.parsing_classes.Var`

	:param indicator_parameter: The indicator parameter to link the variable to
	:type indicator_parameter: class:`ope_anafi.models.IndicatorParameter`
	"""
    variable_split_regex = re.compile(r'([a-zA-Z]*)([0-9]*)')

    split_variable = variable_split_regex.search(variable.name)
    try:
        Variable.objects.get(name=variable.name, numero_compte=split_variable.group(2), type_solde=variable.type_solde,
                             solde=split_variable.group(1), indicator_parameter=indicator_parameter)
    except ObjectDoesNotExist:
        var = VariableSerializer(data={
            "name": variable.name.strip(),
            "numero_compte": split_variable.group(2).strip(),
            "type_solde": variable.type_solde,
            "solde": split_variable.group(1).strip(),
            "indicator_parameter": indicator_parameter.id
        })
        var.is_valid(raise_exception=True)
        var.save()


def extract_indicators_and_variables(tree, indicators, indicator_parameter):
    """
	Recursive function that will create the Variables and retrieve the indicators from a parsed equation (a tuple)

	:param tree: The node to a analyze
	:type tree: tuple, Var, Indic, str, int

	:param indicators: The array containing all the indicators present in the equation
	:type indicators: list

	:param indicator_parameter: The indicator parameter to link the variables to
	:type indicator_parameter: class:`open_anafi.models.IndicatorParameter`

	:return: A tuple of the equation that can be read by the report parser
	:rtype: tuple, str, int
	"""
    if type(tree) is tuple:
        return tree[0], extract_indicators_and_variables(tree[1], indicators,
                                                         indicator_parameter), extract_indicators_and_variables(tree[2],
                                                                                                                indicators,
                                                                                                                indicator_parameter)
    elif type(tree) is Var:
        create_variable(tree, indicator_parameter)
        if tree.minus:
            return "-" + tree.name
        return tree.name
    elif type(tree) is Indic:
        indicators.append(tree.name)

        return_value = tree.name

        if tree.offset is not None:
            return_value = f"@{tree.offset}" + return_value
        if tree.minus:
            return_value = "-" + return_value
        return return_value
    else:
        return tree


def create_parameter(parsed_parameter, formule, indicator):
    """
	Creates an indicator parameter.

	:param parsed_parameter: The dictionary corresponding to a single parameter, parsed by the langage analyzer
	:type parsed_parameter: dict

	:param formule: The original equation (as a string typed by the user)
	:type formule: str

	:param indicator: The indicator to link the parameter to
	:type indicator: class:`open_anafi.models.Indicator`

	:return: The created indicator parameter
	:rtype: class:`open_anafi.models.IndicatorParameter`
	"""
    establishment_number_regexp = re.compile(r'^[0-9]{3}$')

    cal = parsed_parameter.get('CAL')
    type_budget = parsed_parameter.get('typeBudget')
    type_etablissement = parsed_parameter.get('TE')
    tree = parsed_parameter.get('tree')
    year_min = parsed_parameter.get('exmin')
    year_max = parsed_parameter.get('exmax')
    indicators = []

    indicator_parameter = IndicatorParameter.objects.create()

    parsed_tree = extract_indicators_and_variables(tree, indicators, indicator_parameter)

    indicator_parameter.indicator = indicator
    indicator_parameter.readable_equation = parsed_tree
    indicator_parameter.year_min = year_min
    indicator_parameter.year_max = year_max
    indicator_parameter.original_equation = formule
    indicator_parameter.displaying = cal
    indicator_parameter.type_budget = type_budget

    if type_etablissement is not None:
        establishment_string = establishment_number_regexp.match(type_etablissement)
        establishment_number = establishment_number_regexp.match(type_etablissement)

        if establishment_string is not None:
            indicator_parameter.establishment_string = str(establishment_string)
        if establishment_number is not None:
            indicator_parameter.establishment_number = int(establishment_number)

    indicator_parameter.save()
    return indicator_parameter


def check_for_undefined_indicators(formula):
    """
	Runs through an equation tree and checks if every indicator found exists. Raises a `class: django.core.exceptions.ObjectDoesNotExist` exception if not.

	:param formula: The result of a parseEquation() call. Contains all the trees to run through
	:type formula:
	"""
    if type(formula) is Indic:
        Indicator.objects.get(name=formula.name)
    elif type(formula) is tuple:
        # We recursively check left and right child
        check_for_undefined_indicators(formula[1])
        check_for_undefined_indicators(formula[2])


def parse_indicator(indicator_name, indicator_equation, indicator_description, **kwargs):
    """Creates an indicator by parsing its formula.

	:param indicator_name: The name of the indicator
	:type indicator_name: str

	:param indicator_equation: The equation to parse
	:type indicator_equation: str

	:param indicator_description: The description of the indicator
	:type indicator_description: str

	"""
    split_formula_by_exercice = re.compile(
        r'\[RE-MIN:[\s\d#]+\]|\[RE-MAX:[\s\d#]+\]|(\[TB:[\sA-Z0-9]+\])|(\[TE:[\sA-Z0-9]+\])')

    frame = kwargs.get('frame', None)
    user = kwargs.get('user', None)
    public = kwargs.get('public', True)
    libelle = kwargs.get('libelle', None)

    # Parsing de l'indicateur pour récupérer les éléments
    indicator_to_parse = indicator_equation
    try:
        parsed_indicator = parse_equation(indicator_to_parse)
    except Exception as e:
        raise Exception(f"Erreur au niveau de l'indicateur {indicator_name.strip()} - {str(e)}")

    # On vérifie si un des indicateurs présents n'existe pas. On renvoie une exception si c'est le cas.
    try:
        for param in parsed_indicator:
            check_for_undefined_indicators(param['tree'])
    except ObjectDoesNotExist as e:
        raise ContainsUndefinedIndicators(str(e))

    try:
        indic = Indicator.objects.get(name=indicator_name)
        # Keep frames associated with indic so the newly created indicator can stay attached to the other frames
        frames_associated = list(indic.frames.all())
        indic.delete()
    except ObjectDoesNotExist:
        frames_associated = []
        pass

    indic = IndicatorSerializer(data={"name": indicator_name,
                                      "description": indicator_description if indicator_description else '',
                                      "max_depth": 0
                                      })
    indic.is_valid(raise_exception=True)

    try:
        indic.save()
    except IntegrityError as e:
        raise Exception(str(e))

    indic = Indicator.objects.get(name=indicator_name)
    indic.public = public
    if user is not None:
        indic.last_modified_by = User.objects.get(username=user)
    indic.save()

    if libelle is not None:
        indicator_libelle = IndicatorLibelle.objects.create(libelle=libelle, indicator=indic)
        indicator_libelle.save()

    if frame is not None:
        frame.indicators.add(indic)
    # Attach the indicator to the frames previously linked to this indicator.
    for fram in frames_associated :
        fram.indicators.add(indic)

    # Création des paramètres
    original_equations = []
    temp = ''

    split_indicator_to_parse = split_formula_by_exercice.split(indicator_to_parse)

    if len(split_indicator_to_parse) == 1:
        original_equations.append(split_indicator_to_parse[0])
    else:
        for item in split_indicator_to_parse:
            if item is None or item is '':
                if temp != '':
                    original_equations.append(temp)
                    temp = ''
                continue
            temp += item

    for index, param in enumerate(reversed(parsed_indicator)):
        indicator_parameter = create_parameter(param,
                                               original_equations[index],
                                               indic)

        IndicatorParameterTools.calculate_depth(indicator_parameter)

    IndicatorTools.calculate_max_depth(indic)
    for fr in indic.frames.all():
        FrameTools.calculate_depth(fr)


def update_formula(indicator_equation, indic):
    # On split la formule en fonction des exercices de gestion
    indicator_to_parse = indicator_equation
    split_formula_by_exercice = re.compile(
        r'\[RE-MIN:[\s\d#]+\]|\[RE-MAX:[\s\d#]+\]|(\[TB:[\sA-Z0-9]+\])|(\[TE:[\sA-Z0-9]+\])')
    if not indicator_equation.strip(): return
    try:
        parsed_indicator = parse_equation(indicator_to_parse)
    except Exception as e:
        raise Exception(f"Impossible de parser l'equation : " + str(e))
    original_equations = []
    temp = ''

    split_indicator_to_parse = split_formula_by_exercice.split(indicator_to_parse)
    if len(split_indicator_to_parse) == 1:
        original_equations.append(split_indicator_to_parse[0])
    else:
        for item in split_indicator_to_parse:
            if item is None or item is '':
                if temp != '':
                    original_equations.append(temp)
                    temp = ""
                continue
            temp += item

    # TODO : Look for clever way to remove duplicated code
    for index, param in enumerate(reversed(parsed_indicator)):
        indicator_parameter = create_parameter(param,
                                               original_equations[index],
                                               indic)

        # This is for the case where you add a single indicator/parameter

        IndicatorParameterTools.calculate_depth(indicator_parameter)

    IndicatorTools.calculate_max_depth(indic)
    for fr in indic.frames.all():
        FrameTools.calculate_depth(fr)


def parse_excel_file(file_name, institution_types, identifier_types, frame_name, frame_description, nomenclature,
                     model_name=None, user=None):
    """The function that parses an excel file containing multiple indicators and creates them.
	   It will also create a frame and link the indicators to it.
	
	:param file_name: The name of the excel file to load
	:type file_name: str

	:param institution_types: The institutions the frame can be calculated on
	:type institution_types: list of int

	:param identifier_types: The types of identifiers that the frame can be calculated on
	:type identifier_types: list of int

	:param frame_name: The name of the frame to create
	:type frame_name: str

	:param frame_description: The description of the frame to create
	:type frame_description: str

	:param nomenclature: The nomenclature to link the frame to
	:type nomenclature: int

	:param model_name: The name of the model file to link to the frame
	:type model_name: optional, str

	:param user: The user who will be stored in the the created indicators.
	:type user: optional, django.contrib.auth.models.User
	"""

    # On récupère la frame
    frame = FrameTools.get_or_create_frame(frame_name, model_name, frame_description, nomenclature, institution_types,
                                           identifier_types)

    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb[wb.sheetnames[0]]

    q = queue.Queue(maxsize = 0)

    # On parcourt le fichier excel ligne par ligne
    for row in ws.iter_rows(min_row=2):
        try:
            description = row[3].value
        except:
            description = ''
        if row[1].value:
            libelle = row[1].value.strip()

        else:
            libelle = ""
        logger.debug('Queuing ', row[0].value.strip() )
        q.put(ExcelEntry( row[0].value.strip(),  libelle, row[2].value.replace(' ', ''), description))
    counter = 0
    while not q.empty():
        entry_to_parse = q.get()
        if counter > q.qsize()+10: raise Exception('Missing Indicator')
        try:
            parse_indicator(entry_to_parse.name,
                            entry_to_parse.formula,
                            entry_to_parse.description,
                            libelle=entry_to_parse.libelle,
                            public=True,
                            user=user,
                            frame=frame)
            logger.debug(entry_to_parse.name)
            counter =0
        except ContainsUndefinedIndicators :
            logger.debug(entry_to_parse.name, 'to the end of queue')
            q.put(entry_to_parse)
            counter +=1
