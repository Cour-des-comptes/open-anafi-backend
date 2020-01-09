import math
from copy import copy
from datetime import datetime
from os.path import dirname, abspath
from random import randrange

import django.db as db
import numpy
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from open_anafi.lib import SQL_REQUEST_COLLECTIVITE_PER_YEAR
from open_anafi.models import IndicatorLibelle
from open_anafi.models import InstitutionType


class ExcelWriterTools:

    @staticmethod
    def write_pseudo(indicators, model_file_path):
        """
        This function returns the recreated pseudo of a Frame as well as the model associated enriched with the indicators formulas.
        :param indicators:
        :param model_file_path:
        :return: workbooks wb, wb2 who are respectively the recreated pseudo and the model including formulas
        """
        wb = load_workbook(dirname(dirname(abspath(__file__))) + '/templates/pseudoparam_template.xlsx')
        ws = wb.active
        indic_formula_dict = {}
        indic_description_dict = {}
        for r, indic in enumerate(indicators.all()):
            cell = ws['{}{}'.format(get_column_letter(1), r + 2)]
            cell.value = indic.name
            cell = ws['{}{}'.format(get_column_letter(2), r + 2)]
            indicator_libelle = IndicatorLibelle.objects.filter(indicator=indic)
            if len(indicator_libelle) > 0:    cell.set_explicit_value(indicator_libelle[0].libelle, 's')
            cell = ws['{}{}'.format(get_column_letter(3), r + 2)]
            formula = ''
            for idx, parameter in enumerate(reversed(indic.parameters.all())):
                if idx > 0 and (parameter.original_equation.lstrip()[0] not in ['+', '-']):
                    formula += ' + '
                ymin, ymax = parameter.year_min, parameter.year_max
                if not ymin: ymin = '#'
                if not ymax: ymax = '#'
                formula += parameter.original_equation
                if not (ymin == '#' and ymax == '#'): formula += '[RE - MIN: {} ][RE - MAX : {} ]'.format(ymin, ymax)

            cell.value = formula
            indic_formula_dict[indic.name] = formula
            if indic.description: indic_description_dict[indic.name] = indic.description
            cell = ws['{}{}'.format(get_column_letter(4), r + 2)]
            cell.value = indic.description
        wb2 = load_workbook(model_file_path)

        for ws in wb2.worksheets:
            max_row = ws.max_row
            for r in range(max_row):
                cell_i = ws['{}{}'.format(get_column_letter(2), r)]
                cell_f = ws['{}{}'.format(get_column_letter(6), r)]
                cell_d = ws['{}{}'.format(get_column_letter(7), r)]
                if cell_i.value in indic_formula_dict.keys():
                    cell_f.value = indic_formula_dict[cell_i.value]
                if cell_i.value in indic_description_dict.keys():
                    cell_d.value = indic_description_dict[cell_i.value]


        return wb, wb2

    @staticmethod
    def write_report(computed_indicators, ident, type_ident, exs, libsiret, model_name, dps=None, aggregate=False,
                     perimeter=None):
        """
        Writes an open_anafi report into an excel file.

        :param perimeter:
        :param computed_indicators: A list of all the indicators and their values for each year to evaluate
        :type computed_indicators: class:`pandas.DataFrame`

        :param ident: The identifier(s) the report was evaluated on (A concatenation if multiple SIRETS)
        :type ident: str

        :param type_ident: The type of the identifier (SIRET, SIREN, FINESS)
        :type type_ident: str

        :param exs: The list of years the repport was calculated on
        :type exs: list

        :param libsiret: The establishment's label (#TODO check this, not sure)
        :type libsiret: str

        :param model_name: The name of the reference excel file (The model to fill)
        :type model_name: str

        :param dps: dictionary with the state of the account (Definitive or Provisoire)
        :type dps: dict

        :param aggregate: If the report was an aggreggation of multiple identifiers or not
        :type aggregate: bool

        :return: The name of the created Excel file
        :rtype: str
        """

        # put data in excel
        global patterns, siret
        wb = load_workbook('./open_anafi/templates/' + model_name)
        double = Side(border_style="thin", color="000000")
        border = Border(top=double, left=double, right=double, bottom=double)
        font_header = Font(italic=True, color="000000", name='Arial', size=7)
        al_header = Alignment(horizontal="center", vertical="center")
        al_dps = Alignment(horizontal="center", vertical="bottom")
        al_indic = Alignment(horizontal="right", vertical="center")
        sheets = wb.worksheets

        for sheet in sheets:
            col5 = {r: sheet['{}{}'.format(get_column_letter(5), r)].value for r in range(1, sheet.max_row)}
            # Garder le formattage conditionnel de la colonne des totaux.
            rules = copy(sheet.conditional_formatting)

            cond_format_dict = {}
            for r in list(rules._cf_rules.items()):
                space = str(r[0]).split(" ")[1][:-1]
                rule = r[1]
                cond_format_dict[space] = rule

            sheet.conditional_formatting = ConditionalFormattingList()
            for key, rule in cond_format_dict.items():
                if key[0] == 'D':
                    k2 = copy(key)
                    for rul in rule:
                        for i in range(len(exs)):
                            sheet.conditional_formatting.add(k2.replace('D', get_column_letter(4 + i)), rul)

            max_row = sheet.max_row
            c = get_column_letter(2)
            r = 1
            total_type = None
            while r < max_row + 1:
                while sheet['{}{}'.format(c, r)].value is None and r < max_row + 1:
                    r += 1
                    total_type = None

                if total_type is None:
                    type_cell = sheet['{}{}'.format(get_column_letter(5), r - 1)]
                    type_cell.alignment = Alignment(horizontal='center')
                    total_type = type_cell.value if type_cell.value is not None else "unknown"

                indicator_name = sheet['{}{}'.format(c, r)].value
                c = get_column_letter(4)
                cell = sheet['{}{}'.format(c, r)]
                style = copy(cell._style)
                hyperlink = copy(cell.hyperlink)
                number_format = copy(cell.number_format)
                try:
                    for i in range(computed_indicators[indicator_name].shape[0]):
                        cell = sheet['{}{}'.format(c, r)]
                        cell.value = computed_indicators[indicator_name][i]
                        cell._style = style
                        cell.hyperlink = hyperlink
                        cell.number_format = number_format
                        c = get_column_letter(5 + i)

                    current_indicator = computed_indicators[indicator_name]
                    if total_type == "$totalV":

                        n = len(current_indicator)
                        m = exs[-1] - exs[0]
                        if n == 1 or n == 0:
                            tvam = 0
                        else:
                            # solution de secour: SI rapport negatif on calcul pas sinon on calcul
                            ratio = current_indicator[-1] / current_indicator[0]
                            tvam = math.copysign(1.0, ratio) * math.pow(abs(ratio), 1 / (m - 1)) - 1

                        all_same_sign = (-1 * numpy.sign(current_indicator[0]) not in [numpy.sign(current_indicator[0]),
                                                                                       numpy.sign(
                                                                                           current_indicator[n - 1])])
                        copied_font = copy(sheet[f'{chr(ord(c) - 1)}{r}'].font)
                        copied_font.italic = True
                        if col5[r] != 'N' and all_same_sign:
                            sheet['{}{}'.format(c, r)].value = tvam

                        sheet['{}{}'.format(c, r)].font = copied_font
                        sheet['{}{}'.format(c, r)].number_format = '0.0%'
                        sheet['{}{}'.format(c, r)].border = copy(sheet[f'{chr(ord(c) - 1)}{r}'].border)
                        sheet['{}{}'.format(c, r)].fill = copy(sheet[f'{chr(ord(c) - 1)}{r}'].fill)
                        sheet['{}{}'.format(c, r)].alignment = al_indic
                    elif total_type == "$totalM":
                        mean = current_indicator.mean()

                        copied_font = copy(sheet[f'{chr(ord(c) - 1)}{r}'].font)
                        copied_format = copy(sheet[f'{chr(ord(c) - 1)}{r}'].number_format)
                        copied_font.italic = True

                        sheet['{}{}'.format(c, r)].value = mean if not numpy.ma.is_masked(mean) else 0

                        sheet['{}{}'.format(c, r)].font = copied_font
                        sheet['{}{}'.format(c, r)].number_format = copied_format
                        sheet['{}{}'.format(c, r)].border = copy(sheet[f'{chr(ord(c) - 1)}{r}'].border)
                        sheet['{}{}'.format(c, r)].fill = copy(sheet[f'{chr(ord(c) - 1)}{r}'].fill)
                        sheet['{}{}'.format(c, r)].alignment = al_indic
                    elif total_type == "$totalC":

                        copied_font = copy(sheet[f'{chr(ord(c) - 1)}{r}'].font)
                        copied_font.italic = True

                        sheet['{}{}'.format(c, r)].value = numpy.sum(numpy.nan_to_num(current_indicator))
                        sheet['{}{}'.format(c, r)].font = copied_font
                        sheet['{}{}'.format(c, r)].number_format = '#,##0_-'
                        sheet['{}{}'.format(c, r)].border = copy(sheet[f'{chr(ord(c) - 1)}{r}'].border)
                        sheet['{}{}'.format(c, r)].fill = copy(sheet[f'{chr(ord(c) - 1)}{r}'].fill)
                        sheet['{}{}'.format(c, r)].alignment = al_indic
                    else:
                        pass

                except KeyError:
                    pass
                finally:
                    r += 1
                    c = get_column_letter(2)

            c = get_column_letter(4)
            r = 1

            while r < max_row + 1:
                while sheet['{}{}'.format(c, r)].value != '$exges' and r < max_row + 1:
                    r += 1
                    c = get_column_letter(5)
                    total_type = sheet['{}{}'.format(c, r)].value
                    c = get_column_letter(4)
                if r < max_row + 1:
                    cell = sheet['{}{}'.format(c, r)]
                    border = copy(cell.border)
                    fill = copy(cell.fill)
                    font = copy(cell.font)
                    for i in range(len(exs)):
                        cell = sheet['{}{}'.format(c, r)]
                        cell.value = exs[i]
                        cell.border = border
                        cell.fill = fill
                        cell.font = font
                        cell.alignment = Alignment(horizontal='center')
                        c = get_column_letter(5 + i)

                    if total_type == "$totalV":
                        sheet['{}{}'.format(c, r)].value = "Var. annuelle moyenne"
                        sheet['{}{}'.format(c, r)].font = font_header
                        sheet['{}{}'.format(c, r)].border = copy(sheet[f'{chr(ord(c) - 1)}{r}'].border)
                        sheet['{}{}'.format(c, r)].alignment = al_header
                        sheet.column_dimensions[f'{c}'].width = 17
                    elif total_type == "$totalM":
                        sheet['{}{}'.format(c, r)].font = font_header
                        sheet['{}{}'.format(c, r)].border = copy(sheet[f'{chr(ord(c) - 1)}{r}'].border)
                        sheet['{}{}'.format(c, r)].alignment = al_header
                        sheet.column_dimensions[f'{c}'].width = 17
                        sheet['{}{}'.format(c, r)].value = "Moyenne"
                    elif total_type == "$totalC":
                        sheet['{}{}'.format(c, r)].value = "Cumul sur les années"
                        sheet['{}{}'.format(c, r)].font = font_header
                        sheet['{}{}'.format(c, r)].border = copy(sheet[f'{chr(ord(c) - 1)}{r}'].border)
                        sheet['{}{}'.format(c, r)].alignment = al_header
                        sheet.column_dimensions[f'{c}'].width = 17
                    c = get_column_letter(4)
            # if the value is a string
            siret = ident['ident'].tolist()
            if len(siret) == 1:
                patterns = {'$libsiret': f'{libsiret}'.replace('[', '').replace(']', '')[1:-1],
                            '$siret': f'{siret}'.replace('[', '').replace(']', '').replace("'", "").replace(" ",
                                                                                                            "")}  # .replace(" ", "")
            else:
                patterns = {'$libsiret': 'Etablissements multiples',
                            '$siret': 'Etablissements multiples'}
            # if the value is a list
            patterns_list = {'$dp': dps}
            for pattern in patterns:
                coordinates = ExcelWriterTools._look_for_pattern(sheet, pattern)
                for coordinate in coordinates:
                    sheet[coordinate[0]].value = sheet[coordinate[0]].value.replace(pattern, patterns[pattern])
                    sheet[coordinate[0]].border = border
            for pattern_list in patterns_list:
                coordinates = ExcelWriterTools._look_for_pattern(sheet, pattern_list)
                for coordinate in coordinates:
                    col = coordinate[1]['col']
                    row = coordinate[1]['row']
                    for el in patterns_list[pattern_list]:
                        sheet[f'{get_column_letter(col)}{row}'].value = el
                        sheet[f'{get_column_letter(col)}{row}'].border = border
                        sheet[f'{get_column_letter(col)}{row}'].alignment = al_dps
                        sheet[f'{get_column_letter(col)}{row}'].font = copy(
                            sheet[f'{get_column_letter(col - 1)}{row}'].font)
                        col += 1

            coordinates = ExcelWriterTools._look_for_pattern(sheet, '$perimeter')

            if coordinates:
                i = 2
                sheet.delete_rows(coordinates[0][1]['row'])
                ExcelWriterTools.write_perimeter_headers(sheet)
                for r in dataframe_to_rows(ident, index=False, header=False):
                    sheet.cell(row=i, column=1).value = r[0]
                    sheet.cell(row=i, column=2).value = r[1]
                    sheet.cell(row=i, column=3).value = r[2]
                    sheet.cell(row=i, column=4).value = r[3]
                    sheet.cell(row=i, column=5).value = r[4]
                    sheet.cell(row=i, column=6).value = r[5]
                    sheet.cell(row=i, column=7).value = r[6]
                    i += 1

                count = len(list(dataframe_to_rows(ident))) - 1
                tab = Table(displayName="Table1", ref=f"A1:G{count}")

                # Add a default style with striped rows and banded columns
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                sheet.add_table(tab)
                sheet.column_dimensions['A'].width = 23
                sheet.column_dimensions['B'].width = 23
                sheet.column_dimensions['C'].width = 23
                sheet.column_dimensions['D'].width = 23
                sheet.column_dimensions['E'].width = 28
                sheet.column_dimensions['F'].width = 28
                sheet.column_dimensions['G'].width = 23
            # for sheet in sheets:
            #   for merged_cell in sheet.merged_cells.ranges:
            #     ExcelWriterTools.border_maker(sheet, str(merged_cell), border=border)
        if not aggregate:
            siret_name = patterns['$siret'].replace(',', 'et')
        else:
            siret_name = str(randrange(1000, 9999))
        if len(siret_name) > 20:
            siret_name = siret_name[:20]

        ExcelWriterTools.write_perimeter(wb, siret, exs, perimeter)

        report_name = 'FRAME_{}_{}_{}_{}_{}.xlsx'.format(siret_name,
                                                         type_ident,
                                                         exs[0],
                                                         exs[-1],
                                                         datetime.now().strftime("%Y%m%d_%H%M%S"))

        wb.save('{}/reports/{}'.format(dirname(dirname(dirname(abspath(__file__)))), report_name))

        return report_name

    @staticmethod
    def write_perimeter_headers(sheet):
        """
        Writes the headers for the perimeter sheet of the report

        :param sheet: The excel sheet to work on
        :type sheet: class:`openpyxl.Worksheet`
        """
        sheet['A1'] = "Code budget"
        sheet['B1'] = "Nomenclature"
        sheet['C1'] = "Type d'établissement"
        sheet['D1'] = "Exercice"
        sheet['E1'] = "Identifiant"
        sheet['F1'] = "Libellé"
        sheet['G1'] = "Population"

    @staticmethod
    def _look_for_pattern(work_sheet, pattern):
        """
        Will look for a given pattern across an excel work sheet and return the locations.

          :param work_sheet: The excel work sheet to work on
          :type work_sheet: openpyxl.Worksheet

          :param pattern: The pattern to look for
          :type pattern: str

          :return: for a given work_sheet,
            it will return a list of tuple, the first element is a string like 'A1'
            the second one is a dictionary {'col', integer, 'row': integer}
          :rtype: dict
        """
        coordinates = []
        max_row = work_sheet.max_row
        max_col = work_sheet.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                coordinate = f'{get_column_letter(col)}{row}'
                current_value = work_sheet[coordinate].value
                if current_value is not None:
                    # the current_value has to be casted to string to use the following syntax 'pattern in iterable'
                    if pattern in str(current_value):
                        coordinates.append((coordinate, {'col': col, 'row': row}))

        return coordinates

    @staticmethod
    def write_balance(identifiers, identifiers_type, year_min, year_max, libsiret, balances):
        """
        Generate a 'Balance des comptes' report, which is a dump of the data in the accounts

        :param identifiers: The identifier(s) the balance was calculated on (concatenation in case of an aggreggation)
        :type identifiers: str

        :param identifiers_type: The type of the identifiers (SIRET, SIREN, FINESS)
        :type identifiers_type: str

        :param year_min: The lower bound of the range the balance was calculated on
        :type year_min: int

        :param year_max: The lower bound of the range the balance was calculated on
        :type year_max: int

        :param libsiret: The establishment's label (#TODO check this, not sure)
        :type libsiret: str

        :param balances: The dataframe containing all the infos to write
        :type balances: class:`pandas.DataFrame`
        """

        balance_name = 'BALANCE_{}_{}_{}_{}_{}.xlsx'.format(identifiers_type,
                                                            f'{identifiers}'.replace(',', 'et').replace('[',
                                                                                                        '').replace(']',
                                                                                                                    '').replace(
                                                                "'", ""), year_min, year_max,
                                                            datetime.now().strftime("%Y%m%d_%H%M%S"))
        wb = Workbook()
        wb = ExcelWriterTools._write_balance(balances, wb, identifiers, libsiret, year_min, year_max)
        del wb["Sheet"]
        wb.save('{}/reports/{}'.format(dirname(dirname(dirname(abspath(__file__)))), balance_name))

        return balance_name

    @staticmethod
    def _write_balance(balances, wb, ident, libsiret, exmin, exmax):
        """
        Writes the 'Balance des comptes' report

        :param balances: The dataframe containing all the infos to write
        :type balances: class:`pandas.DataFrame`

        :param wb: The openpyxl workbook to work on
        :type wb: class:`openpyxl.Workbook`

        :param ident: The identifier(s) the balance was calculated on. (A concatenation in case of an aggreggation)
        :type ident: str

        :param exmin: The lower bound of the range the balance was calculated on
        :type exmin: int

        :param exmax: The lower bound of the range the balance was calculated on
        :type exmax: int

        :param libsiret: The establishment's label (#TODO check this, not sure)
        :type libsiret: str
        """

        ## Balance for each year in the same excel
        sheet = wb.create_sheet("Balances des comptes")
        # strip_rows(sheet)
        columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
        for column in columns:
            sheet.column_dimensions[column].width = 13
        for i in [2, 4, 6, 8]:
            sheet.row_dimensions[i].height = 30

        double = Side(border_style="thin", color="000000")
        thin = Side(border_style="thin", color="DDDDDD")
        border = Border(top=double, left=double, right=double, bottom=double)
        fill = PatternFill("solid", fgColor="C0C0C0")
        primary = PatternFill("solid", fgColor="F8FBFC")
        secondary = PatternFill("solid", fgColor="FCFDFD")
        font = Font(b=True, color="000000", name='Arial', size=10)
        al = Alignment(horizontal="center", vertical="center")

        ExcelWriterTools.style_range(sheet, 'C2:M2', border=border, fill=fill, font=font, alignment=al)
        sheet['C2'].value = "{} (n° SIRET : {})".format(libsiret, ident)

        ExcelWriterTools.style_range(sheet, 'C4:M4', border=border, fill=fill, font=Font(b=True, name='Arial', size=13),
                                     alignment=al)
        sheet['C4'].value = "{} - {}".format(exmin, exmax)

        ExcelWriterTools.style_range(sheet, 'C6:M6', border=border, fill=fill, font=Font(i=True, name='Arial', size=10),
                                     alignment=al)
        sheet[
            'C6'].value = "* Les opérations budgétaires comprennent les opérations réelles + les opérations d'ordre qui sont détaillées dans la colonne 'Dont opérations budgétaires d'ordre'"

        i += 1
        first = False
        r = 9
        c = 1
        df = balances
        if df is not None:
            for _ in df:
                if c in [1, 2, 3, 4, 5, 6]:
                    c += 1
                    continue
                # sheet['{}{}'.format(get_column_letter(c), r - 1)].value = df[column].sum()
                sheet['{}{}'.format(get_column_letter(c), r - 1)].number_format = '#,##0_-'
                c += 1
            c = 1

            if first:
                for head in list(df.columns.values):
                    sheet['{}{}'.format(get_column_letter(c), r)].value = head
                    c += 1
                c = 1
                r = 9
            for index, row in df.iterrows():
                for value in row.iteritems():
                    sheet['{}{}'.format(get_column_letter(c), r)].value = value[1]
                    sheet['{}{}'.format(get_column_letter(c), r)].number_format = '#,##0_-'
                    if c in [1, 2, 3, 4, 5, 6]:
                        sheet['{}{}'.format(get_column_letter(c), r)].border = Border(top=double, bottom=double,
                                                                                      left=double, right=double)
                        sheet['{}{}'.format(get_column_letter(c), r)].font = Font(b=True, color="000000", name="Arial",
                                                                                  size=9)
                        sheet['{}{}'.format(get_column_letter(c), r)].fill = fill
                        sheet['{}{}'.format(get_column_letter(c), r)].alignment = al
                    else:
                        sheet['{}{}'.format(get_column_letter(c), r)].border = Border(top=thin, bottom=thin, left=thin,
                                                                                      right=thin)
                        sheet['{}{}'.format(get_column_letter(c), r)].font = Font(b=False, color="000000", name="Arial",
                                                                                  size=9)
                        if r % 2 == 1:
                            sheet['{}{}'.format(get_column_letter(c), r)].fill = primary
                        else:
                            sheet['{}{}'.format(get_column_letter(c), r)].fill = secondary
                    c += 1
                c = 1
                r += 1
        ExcelWriterTools.write_headers(sheet)
        return wb

    @staticmethod
    def border_maker(ws, cell_range, border=Border()):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param cell_range:
        :param ws:  Excel worksheet instance
        :type ws: class:`openpyxl.Worksheet`

        :param cell_range: An excel range to style (e.g. A1:F20)
        :type cell_range: str

        :param border: An openpyxl Border
        :type border: class:`openpyxl.styles.Border`
        """

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        rows = ws[cell_range]

        for cell in rows[0]:
            cell.border = cell.border + bottom
        for cell in rows[-1]:
            cell.border = cell.border + top

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = l.border + right

    @staticmethod
    def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param cell_range:
        :param alignment:
        :param ws:  Excel worksheet instance
        :type ws: class:`openpyxl.Worksheet`

        :param cell_range: An excel range to style (e.g. A1:F20)
        :type cell_range: str

        :param border: An openpyxl Border
        :type border: class:`openpyxl.styles.Border`

        :param fill: An openpyxl PatternFill or GradientFill
        :type fill: class:`openpyxl.styles.Fill`

        :param font: An openpyxl Font object
        :type font: class:`openpyxl.styles.Font`
        """

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill

    @staticmethod
    def write_perimeter(wb, ident, exer, perimeter):
        """

        :param wb:
        :param exer:
        :param perimeter:
        :param wb:
        :param ident:
        :param exer:
        :param perimeter:
        :return:
        """
        collectivite_query = SQL_REQUEST_COLLECTIVITE_PER_YEAR.format(
            sirets=f"{ident}".replace('[', '(').replace(']', ')'),
            exercice=f"{list(exer)}".replace('[', '(').replace(']', ')'))

        if perimeter is None:
            with db.connections['cci'].cursor() as cursor:
                cursor.execute(collectivite_query)
                columns = [col[0] for col in cursor.description]
                collectivites = [dict(zip(columns, row)) for row in cursor.fetchall()]

            df = pd.DataFrame(collectivites)
        else:
            df = perimeter
        ws = wb.create_sheet("Périmètre")
        institutions_types = InstitutionType.objects.filter(number__in=list(pd.unique(df['ctype'])))
        types_dict = {str(i.number): i.name for i in institutions_types}

        df.rename(columns={'cbudg': 'Type de budget',
                           'cnome': 'Nomenclature',
                           'ctype': 'Type établissement',
                           'exer': 'Exercice',
                           'ident': 'SIRET',
                           'lbudg': 'Nom',
                           'mpoid': 'Population',
                           'dp': 'Définitif/Provisioire'},
                  inplace=True)
        df['Type établissement'] = df['Type établissement'].apply(lambda x: str(x) + ' - ' + types_dict[str(x)])
        df['Type de budget'] = df['Type de budget'].apply(lambda x: 'BP' if int(x) <= 2 else 'BA')
        df = df.sort_values(by=['Exercice'], ascending=False)
        df = df.reset_index(drop=True)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    @staticmethod
    def write_headers(sheet):
        """
        Writes the headers for a balance des comptes report.

        :param sheet: The excel worksheet to work on
        :type sheet: class:`openpyxl.Worksheet`
        """
        black = Side(border_style="thin", color="000000")
        thin = Side(border_style="thin", color="DDDDDD")
        fill_gray = PatternFill("solid", fgColor="C0C0C0")
        font = Font(b=True, color="000000", name='Arial', size=9)
        al = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = {
            "G": "Balance d'entrée Débit",
            "H": "Balance d'entrée Crédit",
            "I": "Opérations budgétaires Débit*",
            "J": "Opérations budgétaires Crédit*",
            "K": "Dont op. budg. ordre Débit",
            "L": "Dont op. budg. ordre Crédit",
            "M": "Op. ordre non budg. Débit",
            "N": "Op. ordre non budg. Crédit",
            "O": "Annul. Débit",
            "P": "Annul. Crédit",
            "Q": "Op. année Débit",
            "R": "Op. année Crédit",
            "S": "Balance de sortie Débit",
            "T": "Balance de sortie Crédit",
            "U": "Solde Débit",
            "V": "Solde Crédit"
        }

        headers_gray = {
            "A": "Exercice",
            "B": "SIRET",
            "C": "Organisme",
            "D": "Type budget",
            "E": "Nomenclature",
            "F": "Compte"
        }

        for key, value in headers_gray.items():
            sheet[f'{key}8'].value = value
            sheet[f'{key}8'].fill = fill_gray
            sheet[f'{key}8'].font = font
            sheet[f'{key}8'].alignment = al
            sheet[f'{key}8'].border = Border(top=black, bottom=black, left=black, right=black)

        for key, value in headers.items():
            sheet[f'{key}8'].value = value
            sheet[f'{key}8'].font = font
            sheet[f'{key}8'].alignment = al
            sheet[f'{key}8'].border = Border(top=thin, bottom=thin, left=thin, right=thin)
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 28
        sheet.column_dimensions['D'].width = 28
        sheet.row_dimensions[8].height = 40
